#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
arbit_logger.py — Excel trade logger for arbitrage.py
Append-only: one row per triggered opportunity.
Requires: pip install openpyxl
Works standalone; arbitrage.py imports it with try/except so it remains optional.
"""

import datetime
import os

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise ImportError("openpyxl required: pip install openpyxl") from exc


EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "arbit_trades.xlsx")

# (column_name, display_width)
COLUMNS = [
    ("timestamp",         20),
    ("pair",              12),
    ("dry_run",            9),
    ("buy_exchange",      13),
    ("sell_exchange",     13),
    ("buy_price_irr",     16),
    ("sell_price_irr",    16),
    ("amount",            12),
    ("max_amount",        12),
    ("ask_vol",           12),
    ("bid_vol",           12),
    ("liquidity_limited", 18),
    ("gross_pct",         11),
    ("fee_total_pct",     14),
    ("transfer_pct",      14),
    ("irt_fee_pct",       13),
    ("net_pct",           11),
    ("net_irt",           15),
    ("net_usdt",          14),
    ("status",            15),
    ("buy_order_id",      20),
    ("sell_order_id",     20),
    ("error_msg",         42),
]

_STATUS_FILL = {
    "executed":     PatternFill("solid", fgColor="C6EFCE"),   # green
    "dry_run":      PatternFill("solid", fgColor="DDEBF7"),   # light blue
    "balance_fail": PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "both_failed":  PatternFill("solid", fgColor="FFC7CE"),   # red
    "leg_risk":     PatternFill("solid", fgColor="FF0000"),   # bright red
}


class TradeLogger:
    def __init__(self, filepath=EXCEL_FILE):
        self.filepath = filepath
        self._init_file()

    # ── file setup ────────────────────────────

    def _init_file(self):
        if os.path.exists(self.filepath):
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.freeze_panes = "A2"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1F4E79")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, (name, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22
        wb.save(self.filepath)

    # ── public API ────────────────────────────

    def log(self, opp, dry_run, status,
            buy_order_id=None, sell_order_id=None, error_msg=None):
        """
        Append one row to the Excel log.

        Parameters
        ----------
        opp           : opportunity dict from ArbitrageEngine.evaluate()
        dry_run       : bool
        status        : "dry_run" | "executed" | "balance_fail" | "both_failed" | "leg_risk"
        buy_order_id  : str or None
        sell_order_id : str or None
        error_msg     : str or None
        """
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        row_num = ws.max_row + 1

        values = [
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            opp["pair"],
            dry_run,
            opp["buy_from"],
            opp["sell_to"],
            round(opp["buy_price"],    2),
            round(opp["sell_price"],   2),
            opp["amount"],
            opp["max_amount"],
            opp["ask_vol"],
            opp["bid_vol"],
            opp["liquidity_limited"],
            round(opp["gross_pct"],    4),
            round(opp["fee_total"],    4),
            round(opp["transfer_pct"], 4),
            round(opp["irt_fee_pct"],  4),
            round(opp["net_pct"],      4),
            round(opp["net_irt"],      2),
            round(opp["net_usdt"],     6),
            status,
            buy_order_id  or "",
            sell_order_id or "",
            error_msg     or "",
        ]

        row_fill = _STATUS_FILL.get(status)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if row_fill:
                cell.fill = row_fill

        wb.save(self.filepath)


# ─────────────────────────────────────────────
#  STANDALONE TEST
# ─────────────────────────────────────────────

if __name__ == "__main__":
    TEST_FILE = "arbit_trades_test.xlsx"
    logger = TradeLogger(TEST_FILE)

    dummy_opp = {
        "pair":              "USDT/TMN",
        "buy_from":          "nobitex",
        "sell_to":           "wallex",
        "buy_price":         750000,
        "sell_price":        757500,
        "amount":            500.0,
        "max_amount":        500.0,
        "ask_vol":           800.0,
        "bid_vol":           600.0,
        "liquidity_limited": False,
        "gross_pct":         1.0,
        "fee_total":         0.55,
        "transfer_pct":      0.16,
        "irt_fee_pct":       0.014,
        "net_pct":           0.276,
        "net_irt":           10350000,
        "net_usdt":          13.8,
        "market_type":       "IRT",
    }

    logger.log(dummy_opp, dry_run=True,  status="dry_run")
    logger.log(dummy_opp, dry_run=False, status="executed",
               buy_order_id="NB-12345", sell_order_id="WX-67890")
    logger.log(dummy_opp, dry_run=False, status="balance_fail",
               error_msg="USDT at nobitex: need 500.00  have 120.45")
    logger.log(dummy_opp, dry_run=False, status="both_failed",
               error_msg="buy=connection error|sell=timeout")
    logger.log(dummy_opp, dry_run=False, status="leg_risk",
               buy_order_id="NB-99999",
               error_msg="sell_failed: connection timeout")

    print("5 test rows written to %s" % TEST_FILE)
    os.remove(TEST_FILE)
    print("Test file removed. OK.")
